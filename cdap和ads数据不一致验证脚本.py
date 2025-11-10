import json
import logging
import os
import time
import traceback
from logging.handlers import RotatingFileHandler

import openpyxl
import pytz
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker, scoped_session

tz = pytz.timezone("Asia/Kolkata")


class Logger(object):
    @classmethod
    def get(cls, name):
        # Create logger
        logger = logging.getLogger(name)
        logger.setLevel(logging.INFO)

        # Formatter
        fmt = '%(asctime)s - L%(lineno)d - %(threadName)s - %(message)s'
        formatter = logging.Formatter(fmt)

        # Stream based handler
        stream_handler = logging.StreamHandler()
        stream_handler.setLevel(logging.INFO)
        stream_handler.setFormatter(formatter)

        # File based handler
        file_handler = RotatingFileHandler(os.path.basename(__file__) + '.log', maxBytes=1024 * 1024 * 1024,
                                           backupCount=10)
        file_handler.setLevel(logging.INFO)
        file_handler.setFormatter(formatter)

        logger.addHandler(stream_handler)
        logger.addHandler(file_handler)

        return logger


logger = Logger.get('CDAP_ADS_Validation')


class CdapAdsValidation:

    def __init__(self):
        # 读取配置文件
        with open('D:/code-py/pythonProject/db_config.json', 'r') as f:
            # with open('db_config.json', 'r') as f:
            self.db_configs = json.load(f)

        # 初始化数据库连接
        self.init_db_connections()

    def init_db_connections(self):
        """初始化数据库连接"""
        # CDS MySQL连接
        self.cds_engine = create_engine(
            'mysql+mysqlconnector://{user}:{password}@{host}:{port}/{database}'.format(**self.db_configs['cds']),
            pool_size=20, max_overflow=20)
        self.cds_session = scoped_session(sessionmaker(bind=self.cds_engine))

        # CDS PostgreSQL连接
        self.cds_pg_engine = create_engine(
            'postgresql://{user}:{password}@{host}:{port}/{database}'.format(**self.db_configs['cds_pg']),
            pool_size=20, max_overflow=20)
        self.cds_pg_session = scoped_session(sessionmaker(bind=self.cds_pg_engine))

    def query_cdap_base_detail_data(self, channel, dates):
        """查询1：CDAP后台详细数据（history_active_cohort_cost_calculate_trend）"""
        sql = text("""
            SELECT 
              'history_active_cohort_cost_calculate_trend' as table_name,
              hact.dates,
              hact.bdates,
              hact.channel,
              hact.source,
              hact.campaign_id,
              hact.active,
              hact.history_active_offset_days,
              COALESCE(hacfd.channel_threshold_value, 
                (SELECT channel_threshold_value FROM history_active_channel_config 
                 WHERE channel_prefix = 'DEFAULT_CHANNEL_PREFIX')) as threshold_value
            FROM history_active_cohort_cost_calculate_trend hact
            LEFT JOIN history_active_channel_config_detail hacfd ON hact.channel = hacfd.channel 
            WHERE hact.channel = :channel
              AND hact.dates = :dates
              AND hact.dates = hact.bdates
              AND hact.history_active_offset_days > COALESCE(
                hacfd.channel_threshold_value, 
                (SELECT channel_threshold_value FROM history_active_channel_config 
                 WHERE channel_prefix = 'DEFAULT_CHANNEL_PREFIX')
              )
            ORDER BY hact.dates, hact.channel, hact.source, hact.campaign_id, hact.active DESC
        """)

        try:
            session = self.cds_pg_session()
            results = session.execute(sql, {"channel": channel, "dates": dates}).fetchall()
            return results
        except Exception as e:
            logger.error(f"查询CDAP基础详细数据失败: {e}")
            return []
        finally:
            session.close()

    def query_cdap_campaign_detail_data(self, channel, dates):
        """查询2：CDAP Campaign详细数据"""
        sql = text("""
            SELECT 
              'history_active_cohort_cost_calculate_trend_campaign' as table_name,
              hact.dates,
              hact.bdates,
              hact.channel,
              hact.source,
              hact.campaign_id,
              hact.active,
              hact.history_active_offset_days,
              COALESCE(hacfd.channel_threshold_value, 
                (SELECT channel_threshold_value FROM history_active_channel_config 
                 WHERE channel_prefix = 'DEFAULT_CHANNEL_PREFIX')) as threshold_value
            FROM history_active_cohort_cost_calculate_trend_campaign hact
            LEFT JOIN history_active_channel_config_detail hacfd ON hact.channel = hacfd.channel 
            WHERE hact.channel = :channel
              AND hact.dates = :dates
              AND hact.dates = hact.bdates
              AND hact.history_active_offset_days > COALESCE(
                hacfd.channel_threshold_value, 
                (SELECT channel_threshold_value FROM history_active_channel_config 
                 WHERE channel_prefix = 'DEFAULT_CHANNEL_PREFIX')
              )
            ORDER BY hact.dates, hact.channel, hact.source, hact.campaign_id, hact.active DESC
        """)

        try:
            session = self.cds_pg_session()
            results = session.execute(sql, {"channel": channel, "dates": dates}).fetchall()
            return results
        except Exception as e:
            logger.error(f"查询CDAP Campaign详细数据失败: {e}")
            return []
        finally:
            session.close()

    def query_ads_backend_detail_data(self, channel, dates):
        """查询3：ADS广告后台详细数据"""
        sql = text("""
            SELECT 
              'history_active_cohort_cost_calculate_trend_campaign' as table_name,
              hact.id,
              hact.dates,
              hact.bdates,
              hact.channel,
              hact.source,
              hact.campaign_id,
              hact.active,
              hact.history_active_offset_days,
              COALESCE(hacfd.channel_threshold_value, 
                (SELECT channel_threshold_value FROM history_active_channel_config 
                 WHERE channel_prefix = 'DEFAULT_CHANNEL_PREFIX')) as threshold_value
            FROM ad_keywords_campaign a, history_active_cohort_cost_calculate_trend_campaign hact
            LEFT JOIN history_active_channel_config_detail hacfd ON hact.channel = hacfd.channel 
            WHERE hact.channel = :channel
              AND hact.campaign_id = a.campaign_id
              AND hact.dates = :dates
              AND hact.dates = hact.bdates
              AND hact.history_active_offset_days > COALESCE(
                hacfd.channel_threshold_value, 
                (SELECT channel_threshold_value FROM history_active_channel_config 
                 WHERE channel_prefix = 'DEFAULT_CHANNEL_PREFIX')
              )
            ORDER BY hact.dates, hact.channel, hact.source, hact.campaign_id, hact.active DESC
        """)

        try:
            session = self.cds_pg_session()
            
            # 打印执行的SQL和参数
            logger.info("=" * 80)
            logger.info("执行ADS后台数据查询SQL:")
            logger.info("=" * 80)
            logger.info(f"SQL: {sql}")
            logger.info(f"参数: channel={channel}, dates={dates}")
            logger.info("=" * 80)
            
            results = session.execute(sql, {"channel": channel, "dates": dates}).fetchall()
            
            logger.info(f"ADS后台数据查询原始结果数量: {len(results)}")
            
            # 实现与后台代码一致的去重逻辑：通过id字段去重
            existed_ids = set()
            deduplicated_results = []
            
            for row in results:
                row_id = row[1]  # id字段在第2列（索引1）
                if row_id in existed_ids:
                    logger.warning(f"【ADS后台查询】发现重复数据：【id：{row_id}】")
                    continue
                existed_ids.add(row_id)
                # 去掉id字段，保持与其他查询结果格式一致
                deduplicated_row = (row[0],) + row[2:]  # 保留table_name，去掉id，保留其他字段
                deduplicated_results.append(deduplicated_row)
            
            logger.info(f"ADS后台数据去重后结果数量: {len(deduplicated_results)}")
            
            if len(deduplicated_results) == 0:
                logger.warning("ADS后台数据查询返回空结果，可能原因:")
                logger.warning("1. ad_keywords_campaign表中没有对应的campaign_id")
                logger.warning("2. history_active_cohort_cost_calculate_trend_campaign表中没有匹配的数据")
                logger.warning("3. 查询条件过滤掉了所有数据")
                logger.warning("4. 所有数据都是重复数据被过滤掉了")
                return []
            
            # 实现分组聚合和花费计算逻辑
            processed_results = self.process_ads_data_with_cost(deduplicated_results)
            return processed_results
        except Exception as e:
            logger.error(f"查询ADS后台详细数据失败: {e}")
            return []
        finally:
            session.close()

    def process_ads_data_with_cost(self, results):
        """处理ADS数据：保持原始记录数量，相同campaign_id共用花费值"""
        if not results:
            return results
        
        logger.info("开始处理ADS数据花费计算（保持原始记录数量，相同campaign_id共用花费）")
        
        # 先计算所有唯一campaign_id的花费
        campaign_costs = {}
        for row in results:
            campaign_id = row[5]
            dates = row[1]
            
            if campaign_id not in campaign_costs:
                cost_info = self.calculate_campaign_cost_with_details(dates, campaign_id)
                campaign_costs[campaign_id] = cost_info
                logger.info(f"计算Campaign {campaign_id} 花费: {cost_info['cost_usd']} USD")
        
        # 为每条记录添加对应的花费信息
        processed_results = []
        for row in results:
            campaign_id = row[5]
            
            # 创建新行，添加花费字段
            new_row = list(row)
            cost_info = campaign_costs[campaign_id]
            
            # 添加花费相关字段到结果中
            new_row.append(cost_info['cost_usd'])  # 花费USD
            new_row.append(cost_info['original_cost'])  # 原始花费
            new_row.append(cost_info['currency'])  # 原始币种
            new_row.append(cost_info['rate'])  # 汇率
            new_row.append(cost_info['extra_rate'])  # 额外系数
            
            processed_results.append(tuple(new_row))
        
        logger.info(f"处理完成：共{len(processed_results)}条记录，涉及{len(campaign_costs)}个不同的campaign_id")
        return processed_results
    
    def calculate_campaign_cost_with_details(self, dates, campaign_id):
        """计算Campaign的花费并返回详细信息"""
        result = {
            'cost_usd': 0.0,
            'original_cost': 0.0,
            'currency': 'USD',
            'rate': 1.0,
            'extra_rate': 1.0
        }
        
        try:
            # 1. 查询原始花费
            cost_amount = self.query_campaign_cost(dates, campaign_id)
            result['original_cost'] = cost_amount
            
            if cost_amount is None or cost_amount == 0:
                logger.info(f"Campaign {campaign_id} 花费为0或查询失败")
                return result
            
            # 2. 查询项目配置获取pn
            pn = self.get_pn_by_campaign_id(campaign_id)
            if not pn:
                logger.warning(f"无法获取campaign_id {campaign_id} 对应的pn")
                result['cost_usd'] = cost_amount  # 如果无法获取pn，返回原始金额
                return result
            
            # 3. 汇率转换为USD并获取详细信息
            conversion_result = self.currency_to_usd_with_details(dates, cost_amount, pn)
            
            result.update(conversion_result)
            
            logger.info(f"花费计算详情 - Campaign: {campaign_id}")
            logger.info(f"  原始金额: {result['original_cost']} {result['currency']}")
            logger.info(f"  汇率: {result['rate']}")
            logger.info(f"  额外系数: {result['extra_rate']}")
            logger.info(f"  USD金额: {result['cost_usd']}")
            
            return result
            
        except Exception as e:
            logger.error(f"计算Campaign {campaign_id} 花费失败: {e}")
            return result

    def calculate_campaign_cost_usd(self, dates, campaign_id):
        """计算Campaign的花费（USD）- 保持向后兼容"""
        result = self.calculate_campaign_cost_with_details(dates, campaign_id)
        return result['cost_usd']
    
    def query_campaign_cost(self, dates, campaign_id):
        """查询Campaign花费 - 对应 adjustCostRecordService.sumCost"""
        if not campaign_id or campaign_id.strip() == '':
            return 0.0
            
        sql = text("SELECT SUM(cost) FROM adjust_cost_record WHERE dates = :dates AND campaign_id = :campaign_id")
        
        try:
            session = self.cds_pg_session()  # 使用PG连接
            result = session.execute(sql, {"dates": dates, "campaign_id": campaign_id}).fetchone()
            cost = result[0] if result and result[0] is not None else 0.0
            return float(cost)
        except Exception as e:
            logger.error(f"查询Campaign {campaign_id} 花费失败: {e}")
            return 0.0
        finally:
            session.close()
    
    def get_pn_by_campaign_id(self, campaign_id):
        """根据campaign_id获取pn - 简化实现"""
        # 这里需要根据实际业务逻辑实现
        # 可能需要从campaign表或其他表获取pn信息
        # 暂时返回默认值，实际项目中需要完善
        return "IN"  # 默认返回IN，实际需要查询数据库
    
    def query_project_entity(self, pn):
        """查询项目配置 - 对应 projectService.queryEntityByPn"""
        sql = text("""
            SELECT pn, extra_rate, enable, create_time 
            FROM project 
            WHERE pn = :pn AND enable = 1 
            ORDER BY create_time DESC 
            LIMIT 1
        """)
        
        try:
            session = self.cds_session()  # 使用MySQL连接
            result = session.execute(sql, {"pn": pn}).fetchone()
            if result:
                return {
                    'pn': result[0],
                    'extra_rate': float(result[1]) if result[1] is not None else 1.0,
                    'enable': result[2],
                    'create_time': result[3]
                }
            return None
        except Exception as e:
            logger.error(f"查询项目配置失败 pn: {pn}, error: {e}")
            return None
        finally:
            session.close()
    
    def query_rate_entity(self, dates, base, symbols):
        """查询汇率 - 对应 rateService.queryEntityBySymbols"""
        sql = text("""
            SELECT rate 
            FROM rate 
            WHERE dates = :dates AND base = :base AND symbols = :symbols
            ORDER BY created DESC
            LIMIT 1
        """)
        
        try:
            session = self.cds_session()  # 使用MySQL连接
            result = session.execute(sql, {
                "dates": dates, 
                "base": base, 
                "symbols": symbols
            }).fetchone()
            
            if result and result[0] is not None:
                return float(result[0])
            return None
        except Exception as e:
            logger.error(f"查询汇率失败 dates: {dates}, base: {base}, symbols: {symbols}, error: {e}")
            return None
        finally:
            session.close()
    
    def currency_to_usd_with_details(self, dates, source_money, pn):
        """汇率转换为美元并返回详细信息"""
        result = {
            'cost_usd': 0.0,
            'currency': 'USD',
            'rate': 1.0,
            'extra_rate': 1.0
        }
        
        if source_money is None or pn is None:
            logger.warning(f"currency to usd error, source_money: {source_money}, pn: {pn}")
            return result
        
        # 1. 查询项目配置获取额外系数
        project_entity = self.query_project_entity(pn)
        extra_rate = 1.0
        if project_entity:
            extra_rate = project_entity.get('extra_rate', 1.0)
        result['extra_rate'] = extra_rate
        
        # 2. 金额 * 额外系数
        extra_money = source_money * extra_rate
        
        # 3. 获取项目货币配置
        currency = self.get_currency_by_pn(pn)
        if not currency:
            logger.warning(f"无法获取pn {pn} 对应的货币")
            result['cost_usd'] = source_money
            return result
        result['currency'] = currency
        
        # 4. 查询汇率 - base是基础货币，symbols是目标货币
        rate = self.query_rate_entity(dates, "USD", currency)
        if rate is None:
            # 如果查询不到汇率，对INR使用默认值
            if currency.upper() == "INR":
                rate = 78.5
                logger.warning(f"query rate entity null, use default[{rate}], base: USD, symbols: {currency}")
            else:
                logger.warning(f"query rate entity null, base: USD, symbols: {currency}")
                result['cost_usd'] = source_money
                return result
        result['rate'] = rate
        
        # 5. 计算USD金额
        if rate and rate != 0:
            usd_value = extra_money / rate
            result['cost_usd'] = round(usd_value, 2)
        else:
            result['cost_usd'] = source_money
        
        return result

    def currency_to_usd(self, dates, source_money, pn):
        """汇率转换为美元 - 保持向后兼容"""
        result = self.currency_to_usd_with_details(dates, source_money, pn)
        return result['cost_usd']
    
    def get_currency_by_pn(self, pn):
        """根据pn获取货币类型 - 简化实现"""
        # 实际需要查询project_currency_config表
        # 这里简化处理
        if pn == "IN":
            return "INR"
        return "USD"  # 默认USD



    def query_cost_data_issues(self, dates):
        """查询4：花费数据检查"""
        sql = text("""
            SELECT 
              dates,
              channel,
              campaign_id,
              cost,
              CASE 
                WHEN channel IS NULL THEN 'channel_is_null'
                WHEN channel = '' THEN 'channel_is_empty'
                ELSE 'channel_has_value'
              END as channel_status,
              CASE 
                WHEN channel IS NULL OR channel = '' THEN 'problematic_record'
                ELSE 'normal_record'
              END as data_quality
            FROM adjust_cost_record 
            WHERE dates = :dates
              AND (
                (channel IS NULL OR channel = '')
                OR 
                campaign_id IN (
                  SELECT campaign_id
                  FROM adjust_cost_record 
                  WHERE dates = :dates
                  GROUP BY campaign_id
                  HAVING COUNT(DISTINCT COALESCE(channel, '')) > 1
                )
              )
            ORDER BY 
              campaign_id, 
              CASE WHEN channel IS NULL THEN 1 WHEN channel = '' THEN 2 ELSE 3 END,
              channel
        """)

        try:
            session = self.cds_pg_session()
            results = session.execute(sql, {"dates": dates}).fetchall()
            return results
        except Exception as e:
            logger.error(f"查询花费数据问题失败: {e}")
            return []
        finally:
            session.close()


    def validate_single_channel(self, channel, dates):
        """验证单个渠道的数据"""
        logger.info(f"开始验证渠道: {channel}, 日期: {dates}")

        # 执行所有查询
        cdap_base_data = self.query_cdap_base_detail_data(channel, dates)
        cdap_campaign_data = self.query_cdap_campaign_detail_data(channel, dates)
        ads_backend_data = self.query_ads_backend_detail_data(channel, dates)

        # 数据差异比较功能已去掉
        differences = []

        return {
            'channel': channel,
            'dates': dates,
            'cdap_base_data': cdap_base_data,
            'cdap_campaign_data': cdap_campaign_data,
            'ads_backend_data': ads_backend_data,
            'differences': differences
        }

    def export_to_excel(self, validation_results, cost_issues, output_filename):
        """导出结果到Excel"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # 删除默认sheet


        # 2. CDAP基础数据详情
        cdap_base_sheet = wb.create_sheet('CDAP-Roas报表查询结果明细')
        base_headers = [
            '表名', '日期', 'BDates', '渠道', 'Source', 'Campaign ID', '活跃用户数', '历史活跃偏移天数', '阈值'
        ]
        cdap_base_sheet.append(base_headers)

        for result in validation_results:
            for row in result['cdap_base_data']:
                cdap_base_sheet.append(list(row))

        # 3. CDAP Campaign数据详情
        cdap_campaign_sheet = wb.create_sheet('CDAP-Roas条件模拟明细')
        campaign_headers = [
            '表名', '日期', 'BDates', '渠道', 'Source', 'Campaign ID', '活跃用户数', '历史活跃偏移天数', '阈值'
        ]
        cdap_campaign_sheet.append(campaign_headers)

        for result in validation_results:
            for row in result['cdap_campaign_data']:
                cdap_campaign_sheet.append(list(row))

        # 4. ADS后台数据详情
        ads_backend_sheet = wb.create_sheet('ADS后台数据详情')
        ads_headers = [
            '表名', '日期', 'BDates', '渠道', 'Source', 'Campaign ID', '活跃用户数', '历史活跃偏移天数', '阈值', 
            '花费USD', '原始花费', '原始币种', '汇率', '额外系数'
        ]
        ads_backend_sheet.append(ads_headers)

        for result in validation_results:
            for row in result['ads_backend_data']:
                ads_backend_sheet.append(list(row))

        # 6. 花费数据问题
        cost_sheet = wb.create_sheet('花费数据问题')
        cost_headers = ['日期', '渠道', 'Campaign ID', '花费', '渠道状态', '数据质量']
        cost_sheet.append(cost_headers)

        for cost_issue in cost_issues:
            cost_sheet.append(list(cost_issue))


        wb.save(output_filename)
        logger.info(f"验证结果已导出到: {output_filename}")

    def run_validation(self):
        """执行完整的验证流程"""
        try:
            # 获取配置参数
            dates = self.db_configs['params']['date']
            channels = self.db_configs['params']['channels']

            if not dates or not channels:
                raise ValueError('参数错误，请配置date和channels')

            logger.info(f"开始验证，日期: {dates}, 渠道数量: {len(channels)}")

            # 验证所有渠道
            validation_results = []
            for channel in channels:
                result = self.validate_single_channel(channel, dates)
                validation_results.append(result)

            # 查询花费数据问题
            cost_issues = self.query_cost_data_issues(dates)

            # 导出结果
            timestamp = int(time.time())
            output_filename = f'D:/code-py/pythonProject/{dates}-cdap-ads-validation-detailed-{timestamp}.xlsx'
            # output_filename = f'{dates}-cdap-ads-validation-detailed-{timestamp}.xlsx'  # 本地测试用

            self.export_to_excel(validation_results, cost_issues, output_filename)

            # 打印汇总信息
            self.print_summary(validation_results, cost_issues)

        except Exception as e:
            logger.error(f"验证过程出错: {e}")
            traceback.print_exc()
            raise

    def print_summary(self, validation_results, cost_issues):
        """打印验证汇总信息"""
        logger.info("=" * 60)
        logger.info("验证结果汇总")
        logger.info("=" * 60)

        for result in validation_results:
            channel = result['channel']
            dates = result['dates']

            logger.info(f"渠道: {channel}, 日期: {dates}")
            logger.info(f"  CDAP基础数据记录数: {len(result['cdap_base_data'])}")
            logger.info(f"  CDAP Campaign数据记录数: {len(result['cdap_campaign_data'])}")
            logger.info(f"  ADS后台数据记录数: {len(result['ads_backend_data'])}")
            logger.info("-" * 40)

        if cost_issues:
            logger.info(f"发现花费数据问题记录数: {len(cost_issues)}")

        logger.info("验证完成！详细结果请查看Excel文件。")


if __name__ == '__main__':
    logger.info('-------CDAP和ADS数据详细验证开始-------')
    try:
        validator = CdapAdsValidation()
        validator.run_validation()
        logger.info('-------CDAP和ADS数据详细验证完成-------')
    except Exception as e:
        logger.error(f'验证失败: {e}')
        traceback.print_exc()