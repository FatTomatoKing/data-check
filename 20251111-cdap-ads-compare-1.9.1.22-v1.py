import json
import logging
import os
import time
import traceback
from logging.handlers import RotatingFileHandler

import openpyxl
import pytz
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker, scoped_session
from openpyxl.styles import Alignment

tz = pytz.timezone("Asia/Kolkata")


class Logger(object):
    @classmethod
    def get(cls, name):
        # Create logger
        logger = logging.getLogger(name)
        logger.setLevel(logging.INFO)

        # Formatter
        fmt = '%(asctime)s - L%(lineno)d - %(threadName)s - %(message)s'
        formatter = logging.Formatter(fmt)

        # Stream based handler
        stream_handler = logging.StreamHandler()
        stream_handler.setLevel(logging.INFO)
        stream_handler.setFormatter(formatter)

        # File based handler
        file_handler = RotatingFileHandler(os.path.basename(__file__) + '.log', maxBytes=1024 * 1024 * 1024,
                                           backupCount=10)
        file_handler.setLevel(logging.INFO)
        file_handler.setFormatter(formatter)

        logger.addHandler(stream_handler)
        logger.addHandler(file_handler)

        return logger


logger = Logger.get('CDAP_ADS_Validation')


class CdapAdsValidation:

    def __init__(self):
        # 读取配置文件
        # with open('/opt/cds/db_config.json', 'r') as f:
        with open('db_config.json', 'r') as f:  # 本地用
            self.db_configs = json.load(f)

        # 初始化数据库连接
        self.init_db_connections()

    def init_db_connections(self):
        """初始化数据库连接"""
        # CDS MySQL连接
        self.cds_engine = create_engine(
            'mysql+mysqlconnector://{user}:{password}@{host}:{port}/{database}'.format(**self.db_configs['cds']),
            pool_size=20, max_overflow=20)
        self.cds_session = scoped_session(sessionmaker(bind=self.cds_engine))

        # CDS PostgreSQL连接
        self.cds_pg_engine = create_engine(
            'postgresql://{user}:{password}@{host}:{port}/{database}'.format(**self.db_configs['cds_pg']),
            pool_size=20, max_overflow=20)
        self.cds_pg_session = scoped_session(sessionmaker(bind=self.cds_pg_engine))

    def query_cdap_base_detail_data(self, channel, dates):
        """查询1：CDAP后台详细数据（history_active_cohort_cost_calculate_trend）"""
        sql = text("""
            SELECT 
              'history_active_cohort_cost_calculate_trend' as table_name,
              hact.dates,
              hact.bdates,
              hact.channel,
              hact.source,
              hact.campaign_id,
              hact.active,
              hact.history_active_offset_days,
              hact.day_recharge,
              COALESCE(hacfd.channel_threshold_value, 
                (SELECT channel_threshold_value FROM history_active_channel_config 
                 WHERE channel_prefix = 'DEFAULT_CHANNEL_PREFIX')) as threshold_value
            FROM history_active_cohort_cost_calculate_trend hact
            LEFT JOIN history_active_channel_config_detail hacfd ON hact.channel = hacfd.channel 
            WHERE hact.channel = :channel
              AND hact.dates = :dates
              AND hact.dates <= hact.bdates
              AND hact.cohort = 0
              AND hact.history_active_offset_days > COALESCE(
                hacfd.channel_threshold_value, 
                (SELECT channel_threshold_value FROM history_active_channel_config 
                 WHERE channel_prefix = 'DEFAULT_CHANNEL_PREFIX')
              )
            ORDER BY hact.dates, hact.channel, hact.source, hact.campaign_id, hact.active DESC
        """)

        try:
            session = self.cds_pg_session()
            results = session.execute(sql, {"channel": channel, "dates": dates}).fetchall()
            processed_results = self.process_cdap_data_with_cost(results)
            return processed_results
        except Exception as e:
            logger.error(f"查询CDAP基础详细数据失败: {e}")
            return []
        finally:
            session.close()


    def process_cdap_data_with_cost(self, results):
        """处理CDAP数据：按dates+channel+campaign_id分组计算花费"""
        if not results:
            return results

        logger.info("开始处理CDAP数据花费计算（按dates+channel+campaign_id分组）")
        
        # 按 dates + channel + campaign_id 进行分组，用于确定花费分配逻辑
        group_map = {}
        for row in results:
            # row格式: table_name, dates, bdates, channel, source, campaign_id, active, history_active_offset_days, threshold_value
            dates = row[1]
            channel = row[3] 
            campaign_id = row[5]
            key = f"{dates}_{channel}_{campaign_id}"
            
            if key not in group_map:
                group_map[key] = []
            group_map[key].append(row)
        
        logger.info(f"花费分配分组：共{len(group_map)}个分组")
        
        # 为每个分组计算花费（每个分组只计算一次）
        group_costs = {}
        channel_pn_cache = {}  # 缓存channel对应的pn，避免重复查询
        channel_total_cost_cache = {}  # 缓存channel总花费，避免重复查询
        
        for key, group_rows in group_map.items():
            base_row = group_rows[0]  # 使用第一行获取基本信息
            campaign_id = base_row[5]
            dates = base_row[1]
            channel = base_row[3]
            
            # 通过channel查询对应的pn（使用缓存避免重复查询）
            if channel not in channel_pn_cache:
                channel_pn_cache[channel] = self.get_pn_by_channel(channel)
            pn = channel_pn_cache[channel]
            
            # 查询channel总花费（使用缓存避免重复查询）
            if channel not in channel_total_cost_cache:
                channel_total_cost = self.query_channel_total_cost(dates, channel)
                # 转换为USD
                if channel_total_cost > 0 and pn:
                    channel_total_cost_usd_info = self.currency_to_usd_with_details(dates, channel_total_cost, pn)
                    channel_total_cost_cache[channel] = {
                        'cost_usd': channel_total_cost_usd_info['cost_usd'],
                        'original_cost': channel_total_cost,
                        'currency': channel_total_cost_usd_info['currency'],
                        'rate': channel_total_cost_usd_info['rate'],
                        'extra_rate': channel_total_cost_usd_info['extra_rate']
                    }
                else:
                    channel_total_cost_cache[channel] = {
                        'cost_usd': 0.0,
                        'original_cost': 0.0,
                        'currency': None,
                        'rate': None,
                        'extra_rate': None
                    }
            
            # 计算这个分组的花费 - 添加channel条件
            cost_info = self.calculate_campaign_cost_with_channel_details(dates, campaign_id, channel, pn)
            group_costs[key] = cost_info
            logger.info(f"分组 {key} 花费计算: {cost_info['cost_usd']} USD")
        
        # 处理每条明细记录（保持原始记录数量）
        processed_results = []
        for row in results:
            dates = row[1]
            channel = row[3] 
            campaign_id = row[5]
            key = f"{dates}_{channel}_{campaign_id}"
            
            # 判断是否是该分组的第一条记录（用于花费分配）
            group_rows = group_map[key]
            is_first_record = (row == group_rows[0])
            
            # 处理花费信息：只有每个分组的第一条记录显示花费
            if is_first_record:
                # 分组第一条记录：显示实际花费
                cost_info = group_costs[key]
                logger.info(f"分组 {key} 第一条明细记录，分配花费: {cost_info['cost_usd']} USD")
            else:
                # 分组后续记录：花费设置为0，但保持币种信息
                original_cost_info = group_costs[key]
                cost_info = {
                    'cost_usd': 0.0,
                    'original_cost': 0.0,
                    'currency': original_cost_info['currency'],
                    'rate': original_cost_info['rate'],
                    'extra_rate': original_cost_info['extra_rate']
                }
                logger.info(f"分组 {key} 后续明细记录，花费设置为0")
            
            # 处理day_recharge汇率转换 - 每条明细记录都有自己的充值金额
            day_recharge_raw = row[8] if row[8] is not None else 0.0  # day_recharge在第9列（索引8）
            
            # 获取pn用于day_recharge转换（使用缓存）
            if channel not in channel_pn_cache:
                channel_pn_cache[channel] = self.get_pn_by_channel(channel)
            pn = channel_pn_cache[channel]
            
            try:
                day_recharge = float(day_recharge_raw)
            except (ValueError, TypeError):
                logger.warning(f"day_recharge类型转换失败: {day_recharge_raw}, 设置为0")
                day_recharge = 0.0
            
            # 每条明细记录都进行day_recharge转换
            if day_recharge > 0 and pn:
                recharge_conversion = self.currency_to_usd_with_details(dates, day_recharge, pn)
                logger.info(f"明细记录 Campaign {campaign_id} day_recharge转换: {day_recharge} -> {recharge_conversion['cost_usd']} USD")
            else:
                # day_recharge为0时，仍需要获取正确的货币信息保持一致性
                if pn:
                    currency = self.get_currency_by_pn(pn)
                    project_entity = self.query_project_entity(pn)
                    extra_rate = project_entity.get('extra_rate', 1.0) if project_entity else 1.0
                    rate = self.query_rate_entity(dates, "USD", currency) if currency else None
                    
                    recharge_conversion = {
                        'cost_usd': 0.0,
                        'currency': currency,
                        'rate': rate,
                        'extra_rate': extra_rate
                    }
                else:
                    recharge_conversion = {
                        'cost_usd': 0.0,
                        'currency': None,
                        'rate': None,
                        'extra_rate': None
                    }
            
            # 重新构建行数据，添加花费信息和day_recharge信息
            # 原始数据：table_name, dates, bdates, channel, source, campaign_id, active, history_active_offset_days, day_recharge, threshold_value
            # 新顺序：table_name, dates, bdates, channel, source, campaign_id, active, CDAP实际使用, 原始花费（CDAP实际使用）, 花费USD, 原始花费, 日充值美元, 日充值原始币种
            
            reordered_row = []
            reordered_row.extend(row[:7])  # table_name 到 active
            
            # 获取channel总花费信息
            channel_total_info = channel_total_cost_cache[channel]
            reordered_row.append(channel_total_info['cost_usd'])  # CDAP实际使用
            reordered_row.append(channel_total_info['original_cost'])  # 原始花费（CDAP实际使用）
            
            reordered_row.append(cost_info['cost_usd'])  # 花费USD
            reordered_row.append(cost_info['original_cost'])  # 原始花费
            reordered_row.append(recharge_conversion['cost_usd'])  # 日充值美元
            reordered_row.append(day_recharge)  # 日充值原始币种
            
            processed_results.append(tuple(reordered_row))
        
        logger.info(f"明细数据处理完成：共{len(processed_results)}条明细记录，基于{len(group_map)}个分组的花费分配逻辑")
        return processed_results



    def query_ads_backend_detail_data(self, channel, dates):
        """查询3：ADS广告后台详细数据"""
        sql = text("""
            SELECT 
              'history_active_cohort_cost_calculate_trend_campaign' as table_name,
              hact.id,
              hact.dates,
              hact.bdates,
              hact.channel,
              hact.source,
              hact.campaign_id,
              hact.active,
              hact.history_active_offset_days,
              hact.day_recharge,
              hact.pn,
              COALESCE(hacfd.channel_threshold_value, 
                (SELECT channel_threshold_value FROM history_active_channel_config 
                 WHERE channel_prefix = 'DEFAULT_CHANNEL_PREFIX')) as threshold_value
            FROM ad_keywords_campaign a, history_active_cohort_cost_calculate_trend_campaign hact
            LEFT JOIN history_active_channel_config_detail hacfd ON hact.channel = hacfd.channel 
            WHERE hact.channel = :channel
              AND hact.campaign_id = a.campaign_id
              AND hact.dates = :dates
              AND hact.cohort = 0
              AND hact.history_active_offset_days > COALESCE(
                hacfd.channel_threshold_value, 
                (SELECT channel_threshold_value FROM history_active_channel_config 
                 WHERE channel_prefix = 'DEFAULT_CHANNEL_PREFIX')
              )
            ORDER BY hact.dates, hact.channel, hact.source, hact.campaign_id, hact.active DESC
        """)

        try:
            session = self.cds_pg_session()
            
            # 打印执行的SQL和参数
            logger.info("=" * 80)
            logger.info("执行ADS后台数据查询SQL:")
            logger.info("=" * 80)
            logger.info(f"SQL: {sql}")
            logger.info(f"参数: channel={channel}, dates={dates}")
            logger.info("=" * 80)
            
            results = session.execute(sql, {"channel": channel, "dates": dates}).fetchall()
            
            logger.info(f"ADS后台数据查询原始结果数量: {len(results)}")
            
            # 实现与后台代码一致的去重逻辑：通过id字段去重
            existed_ids = set()
            deduplicated_results = []
            
            for row in results:
                row_id = row[1]  # id字段在第2列（索引1）
                if row_id in existed_ids:
                    logger.warning(f"【ADS后台查询】发现重复数据：【id：{row_id}】")
                    continue
                existed_ids.add(row_id)
                # 去掉id字段，保持与其他查询结果格式一致
                deduplicated_row = (row[0],) + row[2:]  # 保留table_name，去掉id，保留其他字段
                deduplicated_results.append(deduplicated_row)
            
            logger.info(f"ADS后台数据去重后结果数量: {len(deduplicated_results)}")
            
            if len(deduplicated_results) == 0:
                logger.warning("ADS后台数据查询返回空结果，可能原因:")
                logger.warning("1. ad_keywords_campaign表中没有对应的campaign_id")
                logger.warning("2. history_active_cohort_cost_calculate_trend_campaign表中没有匹配的数据")
                logger.warning("3. 查询条件过滤掉了所有数据")
                logger.warning("4. 所有数据都是重复数据被过滤掉了")
                return []
            
            # 实现分组聚合和花费计算逻辑
            processed_results = self.process_ads_data_with_cost(deduplicated_results)
            return processed_results
        except Exception as e:
            logger.error(f"查询ADS后台详细数据失败: {e}")
            return []
        finally:
            session.close()

    def process_ads_data_with_cost(self, results):
        """处理ADS数据：保持明细数据，但花费分配参考Java后端分组逻辑"""
        if not results:
            return results
        
        logger.info("开始处理ADS数据花费计算（保持明细数据，花费分配参考Java后端逻辑）")
        
        # 按 dates + channel + campaignId 进行分组，用于确定花费分配逻辑
        group_map = {}
        for row in results:
            # row格式: table_name, dates, bdates, channel, source, campaign_id, active, history_active_offset_days, day_recharge, threshold_value
            dates = row[1]
            channel = row[3] 
            campaign_id = row[5]
            key = f"{dates}_{channel}_{campaign_id}"
            
            if key not in group_map:
                group_map[key] = []
            group_map[key].append(row)
        
        logger.info(f"花费分配分组：共{len(group_map)}个分组")
        
        # 为每个分组计算花费（每个分组只计算一次）
        group_costs_without_channel = {}  # 不带channel的花费
        group_costs_with_channel = {}     # 带channel的花费
        channel_total_cost_cache = {}     # 缓存channel总花费，避免重复查询
        
        for key, group_rows in group_map.items():
            base_row = group_rows[0]  # 使用第一行获取基本信息
            campaign_id = base_row[5]
            dates = base_row[1]
            pn = base_row[9]  # 获取pn字段
            channel = base_row[3]  # 获取channel信息
            
            # 查询channel总花费（使用缓存避免重复查询）
            if channel not in channel_total_cost_cache:
                channel_total_cost = self.query_channel_total_cost(dates, channel)
                # 转换为USD
                if channel_total_cost > 0 and pn:
                    channel_total_cost_usd_info = self.currency_to_usd_with_details(dates, channel_total_cost, pn)
                    channel_total_cost_cache[channel] = {
                        'cost_usd': channel_total_cost_usd_info['cost_usd'],
                        'original_cost': channel_total_cost,
                        'currency': channel_total_cost_usd_info['currency'],
                        'rate': channel_total_cost_usd_info['rate'],
                        'extra_rate': channel_total_cost_usd_info['extra_rate']
                    }
                else:
                    channel_total_cost_cache[channel] = {
                        'cost_usd': 0.0,
                        'original_cost': 0.0,
                        'currency': None,
                        'rate': None,
                        'extra_rate': None
                    }
            
            # 计算这个分组的花费 - 不带channel条件的方法（原有逻辑）
            cost_info_without_channel = self.calculate_campaign_cost_with_details(dates, campaign_id, pn)
            group_costs_without_channel[key] = cost_info_without_channel
            logger.info(f"分组 {key} 花费计算（不带channel）: {cost_info_without_channel['cost_usd']} USD")
            
            # 计算这个分组的花费 - 带channel条件的方法（新增逻辑）
            cost_info_with_channel = self.calculate_campaign_cost_with_channel_details(dates, campaign_id, channel, pn)
            group_costs_with_channel[key] = cost_info_with_channel
            logger.info(f"分组 {key} 花费计算（带channel）: {cost_info_with_channel['cost_usd']} USD")
        
        # 处理每条明细记录（保持原始记录数量）
        processed_results = []
        for row in results:
            dates = row[1]
            channel = row[3] 
            campaign_id = row[5]
            key = f"{dates}_{channel}_{campaign_id}"
            
            # 判断是否是该分组的第一条记录（用于花费分配）
            group_rows = group_map[key]
            is_first_record = (row == group_rows[0])
            
            # 处理花费信息：只有每个分组的第一条记录显示花费
            if is_first_record:
                # 分组第一条记录：显示实际花费
                cost_info_without_channel = group_costs_without_channel[key]
                cost_info_with_channel = group_costs_with_channel[key]
                logger.info(f"分组 {key} 第一条明细记录，分配花费（不带channel）: {cost_info_without_channel['cost_usd']} USD")
                logger.info(f"分组 {key} 第一条明细记录，分配花费（带channel）: {cost_info_with_channel['cost_usd']} USD")
            else:
                # 分组后续记录：花费设置为0，但保持币种信息
                original_cost_info_without_channel = group_costs_without_channel[key]
                original_cost_info_with_channel = group_costs_with_channel[key]
                
                cost_info_without_channel = {
                    'cost_usd': 0.0,
                    'original_cost': 0.0,
                    'currency': original_cost_info_without_channel['currency'],
                    'rate': original_cost_info_without_channel['rate'],
                    'extra_rate': original_cost_info_without_channel['extra_rate']
                }
                
                cost_info_with_channel = {
                    'cost_usd': 0.0,
                    'original_cost': 0.0,
                    'currency': original_cost_info_with_channel['currency'],
                    'rate': original_cost_info_with_channel['rate'],
                    'extra_rate': original_cost_info_with_channel['extra_rate']
                }
                logger.info(f"分组 {key} 后续明细记录，花费设置为0")
            
            # 处理day_recharge汇率转换 - 每条明细记录都有自己的充值金额
            day_recharge_raw = row[8] if row[8] is not None else 0.0
            pn = row[9]  # 直接使用查询出来的pn字段
            
            try:
                day_recharge = float(day_recharge_raw)
            except (ValueError, TypeError):
                logger.warning(f"day_recharge类型转换失败: {day_recharge_raw}, 设置为0")
                day_recharge = 0.0
            
            # 每条明细记录都进行day_recharge转换
            if day_recharge > 0 and pn:
                recharge_conversion = self.currency_to_usd_with_details(dates, day_recharge, pn)
                logger.info(f"明细记录 Campaign {campaign_id} day_recharge转换: {day_recharge} -> {recharge_conversion['cost_usd']} USD")
            else:
                # day_recharge为0时，仍需要获取正确的货币信息保持一致性
                if pn:
                    currency = self.get_currency_by_pn(pn)
                    project_entity = self.query_project_entity(pn)
                    extra_rate = project_entity.get('extra_rate', 1.0) if project_entity else 1.0
                    rate = self.query_rate_entity(dates, "USD", currency) if currency else None
                    
                    recharge_conversion = {
                        'cost_usd': 0.0,
                        'currency': currency,
                        'rate': rate,
                        'extra_rate': extra_rate
                    }
                else:
                    recharge_conversion = {
                        'cost_usd': 0.0,
                        'currency': None,
                        'rate': None,
                        'extra_rate': None
                }
            
            # 重新构建行数据，调整字段顺序（保持明细数据结构，但不输出pn字段）
            # 原始数据：table_name, dates, bdates, channel, source, campaign_id, active, history_active_offset_days, day_recharge, pn, threshold_value
            # 新顺序：table_name, dates, bdates, channel, source, campaign_id, active, 
            #         花费USD（不带channel）, 原始花费（不带channel）, 花费USD（带channel）, 原始花费（带channel）, 日充值美元, 日充值原始币种
            
            reordered_row = []
            reordered_row.extend(row[:7])  # table_name 到 active
            
            # 添加不带channel的花费信息
            reordered_row.append(cost_info_without_channel['cost_usd'])      # 花费USD（不带channel）
            reordered_row.append(cost_info_without_channel['original_cost']) # 原始花费（不带channel）
            
            # 添加带channel的花费信息
            reordered_row.append(cost_info_with_channel['cost_usd'])         # 花费USD（带channel）
            reordered_row.append(cost_info_with_channel['original_cost'])    # 原始花费（带channel）
            
            # 添加充值信息
            reordered_row.append(recharge_conversion['cost_usd'])            # 日充值美元
            reordered_row.append(day_recharge)                              # 日充值原始币种
            
            processed_results.append(tuple(reordered_row))
        
        logger.info(f"明细数据处理完成：共{len(processed_results)}条明细记录，基于{len(group_map)}个分组的花费分配逻辑")
        return processed_results
    
    def calculate_campaign_cost_with_details(self, dates, campaign_id, pn=None):
        """计算Campaign的花费并返回详细信息"""
        result = {
            'cost_usd': 0.0,
            'original_cost': 0.0,
            'currency': None,
            'rate': None,
            'extra_rate': None
        }
        
        try:
            # 1. 查询原始花费
            cost_amount = self.query_campaign_cost(dates, campaign_id)
            result['original_cost'] = cost_amount
            
            if cost_amount is None or cost_amount == 0:
                logger.info(f"Campaign {campaign_id} 花费为0或查询失败")
                return result
            
            # 2. 使用传入的pn参数，如果没有则使用默认值
            if not pn:
                logger.error(f"未提供pn参数")
            
            # 3. 汇率转换为USD并获取详细信息
            conversion_result = self.currency_to_usd_with_details(dates, cost_amount, pn)
            
            result.update(conversion_result)
            
            logger.info(f"花费计算详情 - Campaign: {campaign_id}")
            logger.info(f"  原始金额: {result['original_cost']} {result['currency']}")
            logger.info(f"  汇率: {result['rate']}")
            logger.info(f"  额外系数: {result['extra_rate']}")
            logger.info(f"  USD金额: {result['cost_usd']}")
            
            return result
            
        except Exception as e:
            logger.error(f"计算Campaign {campaign_id} 花费失败: {e}")
            return result

    def calculate_campaign_cost_with_channel_details(self, dates, campaign_id, channel, pn=None):
        """计算Campaign的花费并返回详细信息（带channel条件）"""
        result = {
            'cost_usd': 0.0,
            'original_cost': 0.0,
            'currency': None,
            'rate': None,
            'extra_rate': None
        }
        
        try:
            # 1. 查询原始花费（带channel条件）
            cost_amount = self.query_campaign_cost_by_channel(dates, campaign_id, channel)
            result['original_cost'] = cost_amount
            
            if cost_amount is None or cost_amount == 0:
                logger.info(f"Campaign {campaign_id} Channel {channel} 花费为0或查询失败")
                return result
            
            # 2. 获取pn参数
            if not pn:
                logger.warning(f"Campaign {campaign_id} Channel {channel} 缺少pn参数")
                return result
            
            # 3. 汇率转换为USD并获取详细信息
            conversion_result = self.currency_to_usd_with_details(dates, cost_amount, pn)
            
            result.update(conversion_result)
            
            logger.info(f"花费计算详情 - Campaign: {campaign_id}, Channel: {channel}, PN: {pn}")
            logger.info(f"  原始金额: {result['original_cost']} {result['currency']}")
            logger.info(f"  汇率: {result['rate']}")
            logger.info(f"  额外系数: {result['extra_rate']}")
            logger.info(f"  USD金额: {result['cost_usd']}")
            
            return result
            
        except Exception as e:
            logger.error(f"计算Campaign {campaign_id} Channel {channel} 花费失败: {e}")
            return result

    def calculate_campaign_cost_usd(self, dates, campaign_id, pn=None):
        """计算Campaign的花费（USD）- 保持向后兼容"""
        result = self.calculate_campaign_cost_with_details(dates, campaign_id, pn)
        return result['cost_usd']
    
    def query_campaign_cost_by_channel(self, dates, campaign_id, channel):
        """查询Campaign花费（带channel条件）- 参考Java代码的queryCostByDatas方法"""
        if not campaign_id or campaign_id.strip() == '':
            return 0.0
            
        # 参考Java代码：按dates分组，添加channel条件
        sql = text("""
            SELECT SUM(cost) as cost 
            FROM adjust_cost_record 
            WHERE dates = :dates 
              AND campaign_id = :campaign_id 
              AND channel = :channel
        """)
        
        try:
            session = self.cds_pg_session()  # 使用PG连接
            result = session.execute(sql, {
                "dates": dates, 
                "campaign_id": campaign_id,
                "channel": channel
            }).fetchone()
            cost = result[0] if result and result[0] is not None else 0.0
            return float(cost)
        except Exception as e:
            logger.error(f"查询Campaign {campaign_id} Channel {channel} 花费失败: {e}")
            return 0.0
        finally:
            session.close()

    def query_channel_total_cost(self, dates, channel):
        """查询Channel总花费 - 按channel查询adjust_cost_record表的总花费"""
        if not channel or channel.strip() == '':
            return 0.0
            
        sql = text("SELECT SUM(cost) FROM adjust_cost_record WHERE dates = :dates AND channel = :channel")
        
        try:
            session = self.cds_pg_session()  # 使用PG连接
            result = session.execute(sql, {"dates": dates, "channel": channel}).fetchone()
            cost = result[0] if result and result[0] is not None else 0.0
            return float(cost)
        except Exception as e:
            logger.error(f"查询Channel {channel} 总花费失败: {e}")
            return 0.0
        finally:
            session.close()
    
    def query_campaign_cost(self, dates, campaign_id):
        """查询Campaign花费 - 对应 adjustCostRecordService.sumCost"""
        if not campaign_id or campaign_id.strip() == '':
            return 0.0
            
        sql = text("SELECT SUM(cost) FROM adjust_cost_record WHERE dates = :dates AND campaign_id = :campaign_id")
        
        try:
            session = self.cds_pg_session()  # 使用PG连接
            result = session.execute(sql, {"dates": dates, "campaign_id": campaign_id}).fetchone()
            cost = result[0] if result and result[0] is not None else 0.0
            return float(cost)
        except Exception as e:
            logger.error(f"查询Campaign {campaign_id} 花费失败: {e}")
            return 0.0
        finally:
            session.close()
    
    
    def query_project_entity(self, pn):
        """查询项目配置 - 对应 projectService.queryEntityByPn"""
        sql = text("""
            SELECT pn, extra_rate, enable, create_time 
            FROM project 
            WHERE pn = :pn AND enable = 1 
            ORDER BY create_time DESC 
            LIMIT 1
        """)
        
        try:
            session = self.cds_session()  # 使用MySQL连接
            result = session.execute(sql, {"pn": pn}).fetchone()
            if result:
                return {
                    'pn': result[0],
                    'extra_rate': float(result[1]) if result[1] is not None else 1.0,
                    'enable': result[2],
                    'create_time': result[3]
                }
            return None
        except Exception as e:
            logger.error(f"查询项目配置失败 pn: {pn}, error: {e}")
            return None
        finally:
            session.close()
    
    def query_rate_entity(self, dates, base, symbols):
        """查询汇率 - 对应 rateService.queryEntityBySymbols"""
        sql = text("""
            SELECT rate 
            FROM rate 
            WHERE dates = :dates AND base = :base AND symbols = :symbols
            ORDER BY created DESC
            LIMIT 1
        """)
        
        try:
            session = self.cds_session()  # 使用MySQL连接
            result = session.execute(sql, {
                "dates": dates, 
                "base": base, 
                "symbols": symbols
            }).fetchone()
            
            if result and result[0] is not None:
                return float(result[0])
            return None
        except Exception as e:
            logger.error(f"查询汇率失败 dates: {dates}, base: {base}, symbols: {symbols}, error: {e}")
            return None
        finally:
            session.close()
    
    def currency_to_usd_with_details(self, dates, source_money, pn):
        """汇率转换为美元并返回详细信息"""
        result = {
            'cost_usd': 0.0,
            'currency': None,
            'rate': None,
            'extra_rate': None
        }
        
        if source_money is None or pn is None:
            logger.warning(f"currency to usd error, source_money: {source_money}, pn: {pn}")
            return result
        
        # 1. 查询项目配置获取额外系数
        project_entity = self.query_project_entity(pn)
        extra_rate = 1.0
        if project_entity:
            extra_rate = project_entity.get('extra_rate', 1.0)
        result['extra_rate'] = extra_rate
        
        # 2. 金额 * 额外系数 - 确保类型转换
        try:
            source_money_float = float(source_money)
            extra_rate_float = float(extra_rate)
            extra_money = source_money_float * extra_rate_float
        except (ValueError, TypeError) as e:
            logger.error(f"类型转换失败: source_money={source_money}, extra_rate={extra_rate}, error={e}")
            result['cost_usd'] = 0.0
            return result
        
        # 3. 获取项目货币配置
        currency = self.get_currency_by_pn(pn)
        if not currency:
            logger.warning(f"无法获取pn {pn} 对应的货币")
            return result
        result['currency'] = currency
        
        # 4. 查询汇率 - base是基础货币，symbols是目标货币
        rate = self.query_rate_entity(dates, "USD", currency)
        if rate is None:
            # 如果查询不到汇率，对INR使用默认值
            if currency.upper() == "INR":
                rate = 78.5
                logger.warning(f"query rate entity null, use default[{rate}], base: USD, symbols: {currency}")
            else:
                logger.warning(f"query rate entity null, base: USD, symbols: {currency}")
                return result
        result['rate'] = rate
        
        # 5. 计算USD金额
        if rate and rate != 0:
            try:
                rate_float = float(rate)
                usd_value = extra_money / rate_float
                result['cost_usd'] = round(usd_value, 2)
            except (ValueError, TypeError, ZeroDivisionError) as e:
                logger.error(f"汇率计算失败: extra_money={extra_money}, rate={rate}, error={e}")
                return result
        else:
            logger.warning(f"汇率为0或无效: rate={rate}")
            return result
        
        return result

    def currency_to_usd(self, dates, source_money, pn):
        """汇率转换为美元 - 保持向后兼容"""
        result = self.currency_to_usd_with_details(dates, source_money, pn)
        return result['cost_usd']
    
    def get_pn_by_channel(self, channel_code):
        """根据channel_code查询对应的pn"""
        sql = text("""
            SELECT p.pn, p.extra_rate 
            FROM project p 
            WHERE p.id = (
                SELECT c.project_id 
                FROM channel c 
                WHERE c.id = (
                    SELECT sc.channel_id 
                    FROM sub_channel sc 
                    WHERE sc.channel_code = :channel_code
                )
            )
        """)
        
        try:
            session = self.cds_session()  # 使用MySQL连接
            result = session.execute(sql, {"channel_code": channel_code}).fetchone()
            if result:
                pn = result[0]  # pn字段
                extra_rate = result[1]  # extra_rate字段
                logger.info(f"查询到channel {channel_code} 对应的pn: {pn}, extra_rate: {extra_rate}")
                return pn
            else:
                logger.warning(f"未找到channel {channel_code} 对应的pn配置")
                return None
        except Exception as e:
            logger.error(f"查询channel对应pn失败 channel: {channel_code}, error: {e}")
            return None
        finally:
            session.close()
    
    def get_currency_by_pn(self, pn):
        """根据pn获取货币类型 - 查询project_currency_config表"""
        sql = text("""
            SELECT pn, currency, created 
            FROM project_currency_config 
            WHERE pn = :pn 
            ORDER BY created DESC 
            LIMIT 1
        """)
        
        try:
            session = self.cds_session()  # 使用MySQL连接
            result = session.execute(sql, {"pn": pn}).fetchone()
            if result:
                currency = result[1]  # currency字段
                logger.info(f"查询到pn {pn} 对应的货币: {currency}")
                return currency
            else:
                logger.warning(f"未找到pn {pn} 对应的货币配置")
                return None
        except Exception as e:
            logger.error(f"查询货币配置失败 pn: {pn}, error: {e}")
            return None
        finally:
            session.close()






    def validate_single_channel(self, channel, dates):
        """验证单个渠道的数据"""
        logger.info(f"开始验证渠道: {channel}, 日期: {dates}")

        # 执行查询（删除cdap_campaign_data）
        cdap_base_data = self.query_cdap_base_detail_data(channel, dates)
        ads_backend_data = self.query_ads_backend_detail_data(channel, dates)

        # 数据差异比较功能已去掉
        differences = []

        return {
            'channel': channel,
            'dates': dates,
            'cdap_base_data': cdap_base_data,
            'ads_backend_data': ads_backend_data,
            'differences': differences
        }

    def export_to_excel(self, validation_results, output_filename):
        """导出结果到Excel"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # 删除默认sheet

        # 1. CDAP基础数据详情（带花费信息和day_recharge）
        cdap_base_sheet = wb.create_sheet('CDAP-ROAS趋势(同期群)')
        base_headers = [
            '表名', '注册日期', '行为日期', '渠道', '来源', '广告系列id', '活跃用户数',
            '花费美元（cdap当前逻辑带channel)', '原始花费（cdap当前逻辑带channel）', '花费美元（辅助逻辑带channel）',
            '花费原始币种（辅助逻辑带channel）', '日充值美元', '日充值原始币种'
        ]
        cdap_base_sheet.append(base_headers)

        # 添加数据并处理单元格合并
        current_row = 2  # 从第2行开始（第1行是表头）
        for result in validation_results:
            channel_merge_info = {}  # 记录每个channel的合并信息
            
            # 先添加所有数据
            for row in result['cdap_base_data']:
                cdap_base_sheet.append(list(row))

                # 记录需要合并的channel信息
                channel = row[3]  # 渠道在第4列（索引3）
                cdap_actual_cost = row[7]  # CDAP实际使用在第8列（索引7）
                original_cost = row[8]  # 原始花费（CDAP实际使用）在第9列（索引8）
                
                if channel not in channel_merge_info:
                    channel_merge_info[channel] = {
                        'start_row': current_row,
                        'end_row': current_row,
                        'cdap_actual_cost': cdap_actual_cost,
                        'original_cost': original_cost
                    }
                else:
                    channel_merge_info[channel]['end_row'] = current_row
                
                current_row += 1
            
            # 执行单元格合并
            for channel, info in channel_merge_info.items():
                if info['start_row'] < info['end_row']:  # 只有多行时才合并
                    # 合并CDAP实际使用列（第8列，H列）
                    cdap_base_sheet.merge_cells(f'H{info["start_row"]}:H{info["end_row"]}')
                    # 合并原始花费（CDAP实际使用）列（第9列，I列）
                    cdap_base_sheet.merge_cells(f'I{info["start_row"]}:I{info["end_row"]}')
                    
                    # 设置合并后单元格的值和居中对齐
                    cdap_base_sheet[f'H{info["start_row"]}'].value = info['cdap_actual_cost']
                    cdap_base_sheet[f'I{info["start_row"]}'].value = info['original_cost']
                    
                    # 设置居中对齐
                    alignment = Alignment(horizontal='center', vertical='center')
                    cdap_base_sheet[f'H{info["start_row"]}'].alignment = alignment
                    cdap_base_sheet[f'I{info["start_row"]}'].alignment = alignment

        # 2. ADS后台数据详情
        ads_backend_sheet = wb.create_sheet('ADS-同期群ROAS')
        ads_headers = [
            '表名', '注册日期', '行为日期', '渠道', '来源', '广告系列id', '活跃用户数（ads）',
             '花费美元（ads当前逻辑）', '花费原始币种（ads当前逻辑）',
            '花费美元（ads正确逻辑）', '花费原始币种（ads正确逻辑）','日充值美元（ads）', '日充值原始币种（ads）'
        ]
        ads_backend_sheet.append(ads_headers)

        # 添加数据（ADS后台数据不需要单元格合并）
        for result in validation_results:
            for row in result['ads_backend_data']:
                ads_backend_sheet.append(list(row))

        wb.save(output_filename)
        logger.info(f"验证结果已导出到: {output_filename}")

    def run_validation(self):
        """执行完整的验证流程"""
        try:
            # 获取配置参数
            dates = self.db_configs['params']['date']
            channels = self.db_configs['params']['channels']

            if not dates or not channels:
                raise ValueError('参数错误，请配置date和channels')

            logger.info(f"开始验证，日期: {dates}, 渠道数量: {len(channels)}")

            # 验证所有渠道
            validation_results = []
            for channel in channels:
                result = self.validate_single_channel(channel, dates)
                validation_results.append(result)


            # 导出结果
            timestamp = int(time.time())
            output_filename = f'{dates}-cdap-ads-validation-detailed-{timestamp}.xlsx' # 本地测试用
            #output_filename = f'/opt/cds/datas/{dates}-cdap-ads-validation-detailed-{timestamp}.xlsx'

            self.export_to_excel(validation_results, output_filename)

            # 打印汇总信息
            self.print_summary(validation_results)

        except Exception as e:
            logger.error(f"验证过程出错: {e}")
            traceback.print_exc()
            raise

    def print_summary(self, validation_results):
        """打印验证汇总信息"""
        logger.info("=" * 60)
        logger.info("验证结果汇总")
        logger.info("=" * 60)

        for result in validation_results:
            channel = result['channel']
            dates = result['dates']

            logger.info(f"渠道: {channel}, 日期: {dates}")
            logger.info(f"  CDAP基础数据记录数: {len(result['cdap_base_data'])}")
            logger.info(f"  ADS后台数据记录数: {len(result['ads_backend_data'])}")
            logger.info("-" * 40)

        logger.info("验证完成！详细结果请查看Excel文件。")




if __name__ == '__main__':
    logger.info('-------CDAP和ADS数据详细验证开始-------')
    try:
        validator = CdapAdsValidation()
        validator.run_validation()
        logger.info('-------CDAP和ADS数据详细验证完成-------')
    except Exception as e:
        logger.error(f'验证失败: {e}')
        traceback.print_exc()